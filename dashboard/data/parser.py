"""
Parses the AIS Security Tooling Excel workbook into ToolSummary objects.

Key structural facts (verified by inspection):
- Security Tooling sheet: rows 1-34 contain real data (max_row=35 is safe)
- Col 0 (A): maturity level label on first item of each group
- Col 1 (B): checklist item name
- Col 2 onwards: 14 tool blocks, each 3 columns (Applicable, Status, Comment)
- Tool blocks start at col indices: 2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32, 35, 38, 41
- % Complete sheet: col 1 = Maturity (int), col 4 = % Complete (float 0-1)
"""

import os
import openpyxl

from data.models import (
    ChecklistItem, ToolTask, ToolSummary,
    MATURITY_LABELS, STATUS_COLORS,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Canonical tool names in the order they appear as column blocks (col offset)
TOOL_COL_MAP: dict[str, int] = {
    "XSIAM SIEM":             2,
    "XSIAM XDR":              5,
    "Prisma Access":          8,
    "Firewalls":             11,
    "PKI":                   14,
    "XSIAM Cloud":           17,
    "ServiceNow":            20,
    "XSIAM ASM":             23,
    "IDAM Verify Access":    26,
    "IDAM Verify Directory": 29,
    "IDAM Verify Governance":32,
    "IDAM Verify Privilege": 35,
    "QRadar SIEM":           38,
    "Tenable":               41,
}

# Checklist item rows (0-indexed) mapped to their maturity level.
# Rows that are milestone markers ("Maturity Level X Reached") are excluded.
ITEM_ROW_MAP: list[tuple[int, int]] = [
    # (0-indexed row, maturity_level)
    (3,  1), (4,  1),                           # L1
    (6,  2), (7,  2), (8,  2), (9,  2),         # L2
    (11, 3), (12, 3), (13, 3), (14, 3), (15, 3),# L3
    (17, 4), (18, 4), (19, 4), (20, 4),          # L4
    (22, 5), (23, 5), (24, 5), (25, 5), (26, 5), # L5
    (27, 5), (28, 5), (29, 5), (30, 5), (31, 5), # L5 cont.
]

# Statuses that count as "applicable" work (not N/A)
APPLICABLE_STATUSES = {"Complete", "In Progress", "Pending", "TBC"}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean(value) -> str:
    """Normalise a cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\xa0", " ").strip()


def _normalise_tool_name(raw: str) -> str:
    """Map a raw Excel tool name (may have \n or double spaces) to canonical."""
    cleaned = raw.replace("\n", " ").strip()
    # Collapse multiple spaces
    while "  " in cleaned:
        cleaned = cleaned.replace("  ", " ")
    return cleaned


def _normalise_status(status_raw) -> str:
    """Map raw status cell value to one of our canonical statuses."""
    s = _clean(status_raw)
    mapping = {
        "y": "Complete",
        "n": "N/A",
        "complete": "Complete",
        "in progress": "In Progress",
        "pending": "Pending",
        "tbc": "TBC",
        "n/a": "N/A",
        "": "N/A",
    }
    return mapping.get(s.lower(), s if s else "N/A")


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def get_excel_path() -> str:
    """Return the absolute path to the workbook."""
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(here)
    parent = os.path.dirname(root)
    return os.path.join(parent, "Acceptance Into Support Checklist .xlsm")


def parse_excel(filepath: str) -> list[ToolSummary]:
    """
    Read the workbook and return a list of ToolSummary objects (one per tool).
    """
    wb = openpyxl.load_workbook(
        filepath, read_only=True, keep_vba=False, data_only=True
    )

    # ------------------------------------------------------------------ #
    # 1. Read % Complete sheet for maturity level and pct_complete         #
    # ------------------------------------------------------------------ #
    pct_sheet = wb["% Complete"]
    pct_data: dict[str, tuple[int, float]] = {}  # tool_name -> (maturity, pct)

    for row in pct_sheet.iter_rows(values_only=True, min_row=2, max_row=20):
        tool_raw, maturity, _, _, pct = row[0], row[1], row[2], row[3], row[4]
        if tool_raw is None:
            continue
        tool_name = _normalise_tool_name(str(tool_raw))
        mat = int(maturity) if maturity is not None else 0
        pct_val = float(pct) * 100 if pct is not None else 0.0
        pct_data[tool_name] = (mat, round(pct_val, 1))

    # ------------------------------------------------------------------ #
    # 2. Read Security Tooling sheet for all task data                     #
    # ------------------------------------------------------------------ #
    st_sheet = wb["Security Tooling"]
    rows = list(st_sheet.iter_rows(values_only=True, max_row=35))

    # Build a lookup: canonical tool name -> col start index
    # Verified from row 1 of the sheet but we trust TOOL_COL_MAP
    all_tasks: dict[str, list[ToolTask]] = {t: [] for t in TOOL_COL_MAP}

    for row_idx, maturity_level in ITEM_ROW_MAP:
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        item_name = _clean(row[1]) if len(row) > 1 else ""
        if not item_name:
            continue
        item = ChecklistItem(maturity_level=maturity_level, item_name=item_name)

        for tool_name, col_start in TOOL_COL_MAP.items():
            applicable_raw = row[col_start]     if len(row) > col_start     else None
            status_raw     = row[col_start + 1] if len(row) > col_start + 1 else None
            comment_raw    = row[col_start + 2] if len(row) > col_start + 2 else None

            applicable = _clean(applicable_raw).upper() or "N"
            status     = _normalise_status(status_raw)
            comment    = _clean(comment_raw) or None

            task = ToolTask(
                tool=tool_name,
                item=item,
                applicable=applicable,
                status=status,
                comment=comment,
            )
            all_tasks[tool_name].append(task)

    # ------------------------------------------------------------------ #
    # 3. Assemble ToolSummary objects                                      #
    # ------------------------------------------------------------------ #
    summaries: list[ToolSummary] = []
    for tool_name in TOOL_COL_MAP:
        # Try exact match first, then normalised match
        mat, pct = pct_data.get(tool_name, (None, None))
        if mat is None:
            # Fuzzy fallback: match ignoring case/spaces
            for k, v in pct_data.items():
                if k.lower().replace(" ", "") == tool_name.lower().replace(" ", ""):
                    mat, pct = v
                    break
        if mat is None:
            mat, pct = 0, 0.0

        summary = ToolSummary(
            tool=tool_name,
            current_maturity=mat,
            pct_complete=pct,
            tasks=all_tasks[tool_name],
        )
        summaries.append(summary)

    wb.close()
    return summaries


# ---------------------------------------------------------------------------
# Quick test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    path = get_excel_path()
    print(f"Reading: {path}\n")
    summaries = parse_excel(path)
    print(f"{'Tool':<30} {'Maturity':>8} {'% Done':>8} {'Tasks':>6}")
    print("-" * 58)
    for s in summaries:
        print(f"{s.tool:<30} {s.current_maturity:>8} {s.pct_complete:>7.1f}% {len(s.tasks):>6}")
    total_tasks = sum(len(s.tasks) for s in summaries)
    print(f"\nTotal task records: {total_tasks} (expected 14×25=350)")
